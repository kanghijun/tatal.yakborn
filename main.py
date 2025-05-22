from flask import Flask, render_template_string, render_template, request, send_file, send_from_directory, redirect, url_for
import pandas as pd
from io import StringIO, BytesIO
from datetime import datetime, timedelta
import os
import io as io2
import re
from threading import Lock
from openpyxl import load_workbook

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

YKIHO_DB_PATH = os.path.join(BASE_DIR, "ykiho_db.csv")
TEMPLATE_GAPJI = os.path.join(BASE_DIR, "공급내역보고_갑지_S001-출고.xlsx")
TEMPLATE_EULJI = os.path.join(BASE_DIR, "공급내역보고_을지_S001-출고.xlsx")

if not os.path.exists(YKIHO_DB_PATH):
    pd.DataFrame(columns=["공급받은자사업자번호", "공급받는자요양기관코드", "약국명"]).to_csv(YKIHO_DB_PATH, index=False)

COLUMN_MAP = {
    "곱급받은자사업자번호": "공급받은자사업자번호",
    "공급받은자 사업자번호": "공급받은자사업자번호",
    "곱급받은자요양기관코드": "공급받는자요양기관코드",
    "공급받는자 요양기관코드": "공급받는자요양기관코드",
}

def normalize_columns(df):
    return df.rename(columns={k: v for k, v in COLUMN_MAP.items() if k in df.columns})

def normalize(text):
    return re.sub(r'\s+', '', str(text)).lower()

FILE_STORE = {}
STORE_LOCK = Lock()

def generate_template():
    home_link = "<p><a href='/'>← 메인으로 돌아가기</a></p>"
    db = pd.read_csv(YKIHO_DB_PATH, dtype=str)
    total = db["공급받은자사업자번호"].nunique() if not db.empty else 0
    latest = db.iloc[-1] if not db.empty else {}
    latest_info = f"{latest.get('공급받은자사업자번호', '')} ({latest.get('약국명', '')})"
    return home_link + f"""
<!doctype html>
<title>공급내역 자동 변환기</title>
<h2>① 공급내역보고 갑지</h2>
<form method="post">
  <textarea name="data_gapji" rows=10 cols=80 placeholder="▶ 갑지 S01 데이터 복사 후 붙여넣기"></textarea><br>
  <input type="submit" value="갑지 변환 실행">
</form>

<h2>② 공급내역보고 을지</h2>
<form method="post" action="/eulji">
  <textarea name="data_eulji" rows=10 cols=80 placeholder="▶ 을지 S01 데이터 복사 후 붙여넣기"></textarea><br>
  <input type="submit" value="을지 변환 실행">
</form>

<h2>③ DB 업데이트</h2>
<p>총 등록: <b>{total}개</b>, 마지막: <b>{latest_info}</b></p>
<form method="post" action="/update-db">
  <textarea name="dbdata" rows=5 cols=80 placeholder="약국명[TAB]사업자번호[TAB]요양기관코드"></textarea><br>
  <input type="submit" value="DB 업데이트">
</form>
"""

@app.route("/")
def home():
    return """
    <h2>약국총괄사업본부 업무 간소화</h2>
    <ul>
      <li><a href='/supply'>공급내역보고</a></li>
      <li><a href='/shortage'>품절약 배분</a></li>
    </ul>
    """

@app.route("/supply", methods=["GET", "POST"])
def supply():
    if request.method == "POST" and "data_gapji" in request.form:
        return process_gapji(request.form["data_gapji"])
    return render_template_string(generate_template())

def process_gapji(raw):
    lines = raw.splitlines()
    if len(lines) < 3:
        return "<h3>❌ 데이터 부족: 최소 3줄 필요</h3><a href='/supply'>돌아가기</a>"
    df = pd.read_csv(StringIO("\n".join(lines[2:])), sep="\t", dtype=str)
    df = normalize_columns(df)
    for col in ["공급받은자사업자번호", "공급받는자요양기관코드", "연번", "공급구분"]:
        if col not in df.columns:
            return f"<h3>❌ 컬럼 누락: {col}</h3><a href='/supply'>돌아가기</a>"
    db = pd.read_csv(YKIHO_DB_PATH, dtype=str)
    df = pd.merge(df, db, on="공급받은자사업자번호", how="left", suffixes=("", "_db"))
    df["공급받는자요양기관코드"] = df["공급받는자요양기관코드_db"].fillna(df["공급받는자요양기관코드"])
    df = df.sort_values("연번")
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y%m%d')
    outputs = {}
    for code, suffix, template in [("1","출고", TEMPLATE_GAPJI), ("2","반품", TEMPLATE_GAPJI.replace("출고","반품"))]:
        part = df[df["공급구분"]==code].copy()
        if not part.empty:
            wb = load_workbook(template)
            ws = wb['보고서식']
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=part.shape[1]):
                for cell in row:
                    cell.value = None
            for r_idx, row in enumerate(part.itertuples(index=False), start=2):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            name = f"{yesterday}_S01_공급내역보고_갑지_S001-{suffix}.xlsx"
            outputs[name] = buf.read()
    with STORE_LOCK:
        FILE_STORE['files'] = outputs
    links = ''.join([f"<li><a href='/download?name={name}'>{name}</a></li>" for name in outputs.keys()])
    return f"<h3>갑지 변환 완료!</h3><ul>{links}</ul><a href='/supply'>돌아가기</a>"

def process_eulji(raw):
    lines = raw.splitlines()
    if len(lines) < 4:
        return "<h3>❌ 데이터 부족</h3><a href='/supply'>돌아가기</a>"
    df = pd.read_csv(StringIO("\n".join(lines[3:])), sep="\t", header=None, dtype=str)
    df.iloc[:,0] = pd.to_numeric(df.iloc[:,0], errors='coerce')
    df.iloc[:,4] = pd.to_numeric(df.iloc[:,4], errors='coerce')
    df = df.sort_values(0)
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y%m%d')
    outputs = {}
    for code, suffix, template in [("1","출고", TEMPLATE_EULJI), ("2","반품", TEMPLATE_EULJI.replace("출고","반품"))]:
        part = df[df.iloc[:,5]==code].copy()
        if not part.empty:
            part.drop(columns=5, inplace=True)
            wb = load_workbook(template)
            ws = wb['보고서식']
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=part.shape[1]):
                for cell in row:
                    cell.value = None
            for r_idx, row in enumerate(part.itertuples(index=False), start=2):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            name = f"{yesterday}_S01_공급내역보고_을지_S001-{suffix}.xlsx"
            outputs[name] = buf.read()
    with STORE_LOCK:
        FILE_STORE['files'] = outputs
    links = ''.join([f"<li><a href='/download?name={name}'>{name}</a></li>" for name in outputs.keys()])
    return f"<h3>을지 변환 완료!</h3><ul>{links}</ul><a href='/supply'>돌아가기</a>"

@app.route("/eulji", methods=["POST"])
def eulji():
    return process_eulji(request.form.get("data_eulji", ""))

@app.route("/download")
def download_file():
    name = request.args.get('name')
    with STORE_LOCK:
        files = FILE_STORE.get('files', {})
    if name not in files:
        return "<h3>❌ 파일을 찾을 수 없습니다.</h3><a href='/supply'>돌아가기</a>"
    buf = BytesIO(files[name])
    return send_file(buf, download_name=name, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route("/shortage", methods=['GET', 'POST'])
def shortage():
    if request.method == 'POST':
        if 'stock_pasted' in request.form and request.form['stock_pasted'].strip():
            stock_df = pd.read_csv(io2.StringIO(request.form['stock_pasted']), sep='\t')
        else:
            stock_file = request.files['stock_file']
            stock_df = pd.read_excel(stock_file)

        if 'request_pasted' in request.form and request.form['request_pasted'].strip():
            request_df = pd.read_csv(io2.StringIO(request.form['request_pasted']), sep='\t')
        else:
            request_file = request.files['request_file']
            request_df = pd.read_excel(request_file)

        stock_df["약품명_정규화"] = stock_df["약품명"].apply(normalize)

        result_rows = []
        max_cols = 0

        for _, row in request_df.iterrows():
            center = str(row.get("센터")).strip()
            pharmacy = str(row.get("약국명")).strip()
            pharmacist = str(row.get("약사명")).strip()
            raw = str(row.get("요청 품절약")).strip()
            drugs = [d.strip() for d in raw.split(',') if d.strip()]
            max_cols = max(max_cols, len(drugs))

            row_data = {
                "물류센터": center,
                "약국명": pharmacy,
                "약사명": pharmacist,
            }

            for idx, drug in enumerate(drugs, 1):
                drug_clean = normalize(drug)
                matched = stock_df[
                    (stock_df["센터"].astype(str).str.strip() == center) &
                    (stock_df["약품명_정규화"] == drug_clean)
                ]
                try:
                    재고 = str(int(float(matched["재고"].iloc[0]))) if not matched.empty else "0"
                except:
                    재고 = "0"
                row_data[f"의약품{idx}명"] = drug
                row_data[f"의약품{idx}재고"] = 재고
                row_data[f"__빈칸{idx}__"] = ""

            result_rows.append(row_data)

        result_df = pd.DataFrame(result_rows)

        base_cols = ["물류센터", "약국명", "약사명"]
        dynamic_cols = []
        for i in range(1, max_cols + 1):
            dynamic_cols.append(f"의약품{i}명")
            dynamic_cols.append(f"의약품{i}재고")
            dynamic_cols.append(f"__빈칸{i}__")
        all_columns = base_cols + dynamic_cols
        result_df = result_df.reindex(columns=all_columns)

        filename = f"result_{datetime.today().strftime('%Y%m%d_%H%M%S')}.xlsx"
        result_path = os.path.join(UPLOAD_FOLDER, filename)

        with pd.ExcelWriter(result_path, engine='openpyxl') as writer:
            for center_name, group in result_df.groupby("물류센터"):
                request_cols = [col for col in group.columns if col.startswith("의약품") and col.endswith("명")]
                group["요청개수"] = group[request_cols].notna().sum(axis=1)
                group_sorted = group.sort_values(by="요청개수", ascending=False).drop(columns=["요청개수"])
                group_sorted.to_excel(writer, sheet_name=center_name[:31], index=False)

        return render_template('result.html', table=result_df.to_html(classes='data'), download_link=f'/download-shortage/{filename}')

    return render_template('index.html')

@app.route('/download-shortage/<filename>')
def download_shortage_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

@app.route("/update-db", methods=["POST"])
def update_db():
    raw_db = request.form["dbdata"].strip()
    new = pd.read_csv(StringIO(raw_db), sep="\t", header=None, dtype=str)
    if new.shape[1] < 3:
        return "<h3>❌ 입력 오류</h3><a href='/supply'>돌아가기</a>"
    new.columns = ["약국명", "공급받은자사업자번호", "공급받는자요양기관코드"]
    db = pd.read_csv(YKIHO_DB_PATH, dtype=str)
    combined = pd.concat([db, new], ignore_index=True).drop_duplicates("공급받은자사업자번호", keep="last")
    combined.to_csv(YKIHO_DB_PATH, index=False)
    return redirect("/supply")

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=3000)
